import os
from tkinter import *
from tkinter import messagebox
# from fontTools.ttLib import ttFont
import sqlite3
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
import random
regex = re.compile("[@_!#$%^&*()<>?/|}'\"{~:=;`+,-]")
# from PDFWriter import PDFWriter
days = datetime.now()
year = days.year
last_year = year - 1
month = ['months', f'January {year}', f'February {year}', f'March {year}', f'April {year}', f'May {year}',
         f'June {year}', f'July {year} ', f'August {year}', f'September {year}', f'October {year}',
         f'November{year}', f'December{year}']
int1 = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '0']
table = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
time = ['First Hour', 'Second Hour', 'Third Hour', 'Fourth Hour', 'Fifth Hour', 'Sixth Hour', 'Seventh Hour',
        'Eight Hour', 'Nine Hour', 'Tenth Hour']
orange = PatternFill(patternType='solid', fgColor='D75F00')
red = PatternFill(patternType='solid', fgColor='C00000')
yellow = PatternFill(patternType='solid', fgColor='FFFF00')
green = PatternFill(patternType='solid', fgColor='006633', bgColor='9ACD32')
white = PatternFill(patternType=None)
super_green = PatternFill(patternType='solid', fgColor='00C000')
side1 = Side(border_style='thin', color='9ACD32')
style = Border(left=side1, right=side1, top=side1, bottom=side1)
side = Side(border_style=None)
no_border = Border(left=side, right=side, top=side, bottom=side)
directory = "__Board__"
directory1 = "__Institutes__"
directory2 = '__Section__'
parent_dir = os.getcwd()
college_db_folder = os.path.join(parent_dir, directory)
college_dbpath = os.path.join(college_db_folder, "Board.db")
class_folder = os.path.join(parent_dir, directory1)
section_folder = os.path.join(parent_dir, directory2)
excel_folder = os.path.join(parent_dir, '__Excel__')
background = os.path.join(parent_dir, '__background__')
background_image = os.path.join(background, 'Ullen Aiyya Background.jpg')
class_situation = ['Usual Hour', 'Exam Hour', 'Lab Hour']
constant = "-------------------------------- Select Institution --------------------------------"
constant_class = '----------------------------------- Select Class -----------------------------------'


def click(name, eve, clicked):
    if eve:
        name.configure(state=NORMAL)
        name.delete(0, END)
        name.unbind('<Button>', clicked)
        name['fg'] = 'black'


def hour_data(class_db):
    column = sqlite3.connect(class_db)
    con = column.cursor()
    con.execute("SELECT number FROM  HOUR")
    column_table = table[:[ink[0] for ink in con.fetchall()][0]]
    column.commit()
    column.close()
    return column_table


def students_names(class_db):
    student_names = sqlite3.connect(class_db)
    con = student_names.cursor()
    con.execute("SELECT rowid,name FROM STUDENT")
    names = list(map(lambda x: x[1],  con.fetchall()))
    student_names.commit()
    student_names.close()
    return names


def total_strength(class_db):
    strength = sqlite3.connect(class_db)
    con = strength.cursor()
    my_class = con.execute('SELECT COUNT(*) FROM STUDENT').fetchone()[0]
    strength.commit()
    strength.close()
    return my_class


def name_quality_checker(name):
    def number_checker(given):
        for i in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '0']:
            if i in given:
                return False
        else:
            return True
    if regex.search(name) is None and ']' not in name and '[' not in name and '[]' not in name and '"' not in name and \
            number_checker(name) and '\\' not in name:

        return True
    else:
        return False


def board_deleted(school, _class):
    if os.path.exists(school) is False and os.path.exists(college_dbpath) is True:
        messagebox.showerror(title="Remainder",
                             message=f'Hey u Deleted {_class} Data Base ')
    if os.path.exists(school) is True and os.path.exists(college_dbpath) is False:
        messagebox.showerror(title="Remainder",
                             message=f'Hey u Deleted Board Data Base ')
    if os.path.exists(school) is False and os.path.exists(college_dbpath) is False:
        messagebox.showerror(title="Remainder",
                             message=f'Hey u Deleted Board and {_class} Data Base ')


def final_check(class_name, class_db, close_window, database, institution_name):
    for q in range(1):
        if os.path.exists(class_db) is False and os.path.exists(college_dbpath) is False and os.path.exists(
                database) is False:
            messagebox.showerror(title="Remainder",
                                 message=f'Hey u Deleted Three Data Base ')
        if os.path.exists(class_db) is False and os.path.exists(college_dbpath) is False and os.path.exists(database):
            messagebox.showerror(title="Remainder",
                                 message=f'Hey u Deleted Board and {class_name} Data Base ')
        if os.path.exists(class_db) and os.path.exists(college_dbpath) is False and os.path.exists(database) is False:
            messagebox.showerror(title="Remainder",
                                 message=f'Hey u Deleted  Board and {database} Data Base ')
        if os.path.exists(class_db) and os.path.exists(college_dbpath) is False and os.path.exists(database):
            messagebox.showerror(title="Remainder",
                                 message=f'Hey u Deleted Board Data Base ')
        if os.path.exists(class_db) is False and os.path.exists(college_dbpath) and os.path.exists(database) is False:
            messagebox.showerror(title="Remainder",
                                 message=f'Hey u Deleted {institution_name} and {class_name} Data Base ')
        if os.path.exists(class_db) is False and os.path.exists(college_dbpath) and os.path.exists(database):
            messagebox.showerror(title="Remainder",
                                 message=f'Hey u Deleted {class_name} Data Base ')
        if os.path.exists(class_db) and os.path.exists(college_dbpath) and os.path.exists(database) is False:
            messagebox.showerror(title="Remainder",
                                 message=f'Hey u Deleted {institution_name} Data Base ')
    else:
        close_window.destroy()


def window(entry_window):
    app_width = 1000
    app_height = 563
    screen_width = entry_window.winfo_screenwidth()
    screen_height = entry_window.winfo_screenheight()
    x = (screen_width / 2) - (app_width / 2)
    y = (screen_height / 2) - (app_height / 2)
    entry_window.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
    entry_window.maxsize(width=app_width, height=app_height)
    entry_window.resizable(False, False)
    entry_window.iconbitmap('iron.ico')
    entry_window.title(string="Ullen Aiyya")


def excel_institution():
    for _ in range(1):
        while True:
            if os.path.exists(excel_folder):
                temp_folder = os.path.join(parent_dir, '_temp_')
                try:
                    os.rename(excel_folder, temp_folder)
                except OSError:
                    root1 = Tk()
                    root1.withdraw()
                    messagebox.showinfo("Reminder",
                                        'Kindly close your Folder')
                    root1.destroy()
                else:
                    os.rename(temp_folder, excel_folder)
                    break
            else:
                break
    else:
        if os.path.exists(excel_folder):
            for institution in os.listdir(excel_folder):
                original_institution = os.path.join(excel_folder, institution)
                rename_institution = os.path.join(excel_folder, f'{institution} #Deleted {random.random()}')
                for section in os.listdir(original_institution):
                    original_section = os.path.join(original_institution, section)
                    rename_section = os.path.join(original_institution,
                                                  f'{section} #Deleted {random.random()}')
                    if '#Deleted' not in section:
                        os.rename(original_section, rename_section)
                else:
                    if '#Deleted' not in institution:
                        os.rename(original_institution, rename_institution)


def remove_white_space(name):
    return name.replace(" ", "")


def total_check():
    def college_delete():
        from_board1 = sqlite3.connect(college_dbpath)
        details = from_board1.cursor()
        details.execute("SELECT * FROM COLLEGE")
        college_ = details.fetchall()
        for ken in college_:
            details.execute("DELETE FROM COLLEGE WHERE rowid= " + str(college_.index(ken) + 1))
        from_board1.commit()
        from_board1.close()
        excel_institution()

    def section_deleted():
        for class_db in os.listdir(class_folder):
            if class_db.endswith(".db"):
                section_present = os.path.join(class_folder, class_db)
                if os.path.exists(section_present):
                    ins_name = class_db[:-3]
                    open_institution_check = sqlite3.connect(section_present)
                    pen1 = open_institution_check.cursor()
                    pen1.execute("SELECT * FROM classes")
                    section_data = pen1.fetchall()
                    for _name in section_data:
                        section_name = _name[0]
                        for _ in range(1):
                            while True:
                                if os.path.exists(excel_folder):
                                    temp = os.path.join(parent_dir, '_temp_')
                                    try:
                                        os.rename(excel_folder, temp)
                                    except OSError:
                                        root = Tk()
                                        root.withdraw()
                                        messagebox.showinfo("Reminder",
                                                            'Kindly close your Folder')
                                        root.destroy()
                                        continue
                                    else:
                                        os.rename(temp_folder, excel_folder)
                                        break
                                else:
                                    break
                        else:
                            if os.path.exists(excel_folder):
                                for excel_name in os.listdir(excel_folder):
                                    if remove_white_space(ins_name) == remove_white_space(
                                            excel_name):
                                        original_ins = os.path.join(excel_folder, excel_name)
                                        for section_ in os.listdir(original_ins):
                                            if remove_white_space(section_name) == remove_white_space(section_):
                                                section_org = os.path.join(original_ins, section_)
                                                section_rename = os.path.join(original_ins,
                                                                              f'{section_name} #Deleted '
                                                                              f'{random.random()}')
                                                os.rename(section_org, section_rename)
                            pen1.execute("DELETE FROM classes WHERE rowid= " + str(
                                section_data.index(_name) + 1))
                    else:
                        open_institution_check.commit()
                        open_institution_check.close()

        # if college_db_folder is present

    for ic in range(1):
        if os.path.exists(college_db_folder) and os.path.exists(college_dbpath):
            # If board is present in the  college_db_folder
            from_board = sqlite3.connect(college_dbpath)
            details1 = from_board.cursor()
            details1.execute("SELECT * FROM COLLEGE")
            # if class_folder is present
            if os.path.exists(class_folder):
                for two in details1.fetchall():
                    for name in two:
                        _id_ = []
                        institution_database = os.path.join(class_folder, f'{name}.db')
                        # if class_folder is present and single institution data base also present
                        if os.path.exists(institution_database):
                            # if class_folder is present and then check the section folder is exist
                            if os.path.exists(section_folder):
                                open_class = sqlite3.connect(institution_database)
                                cursor = open_class.cursor()
                                cursor.execute("SELECT * FROM classes")
                                for section in cursor.fetchall():
                                    for section_na in section:
                                        class_delete = os.path.join(section_folder, f'{section_na} {name}.db')
                                        if os.path.exists(class_delete) is False:
                                            class_rowid = []
                                            for i1 in range(1):
                                                for i in range(1):
                                                    while True:
                                                        if os.path.exists(excel_folder) and os.path.exists(
                                                                os.path.join(os.path.join(excel_folder, name),
                                                                             section_na)):
                                                            temp_folder = os.path.join(parent_dir, '_temp_')
                                                            try:
                                                                os.rename(excel_folder, temp_folder)
                                                            except OSError:
                                                                root1 = Tk()
                                                                root1.withdraw()
                                                                messagebox.showinfo("Reminder",
                                                                                    'Kindly close your Folder')
                                                                root1.destroy()
                                                                continue
                                                            else:
                                                                os.rename(temp_folder, excel_folder)
                                                                break
                                                        else:
                                                            break
                                                else:
                                                    original_institute = os.path.join(excel_folder, name)
                                                    original_class = os.path.join(os.path.join(excel_folder, name),
                                                                                  section_na)
                                                    if os.path.exists(original_class):
                                                        rename_class = os.path.join(
                                                            original_institute,
                                                            f'{section_na} #Deleted {random.random()}')
                                                        os.rename(original_class, rename_class)
                                            else:
                                                class_rowid.append(section_na)
                                                cursor.execute("SELECT rowid FROM classes WHERE name = ?",
                                                               class_rowid)
                                                rowid = cursor.fetchone()[0]
                                                cursor.execute("DELETE FROM classes WHERE rowid= " + str(rowid))
                                else:
                                    open_class.commit()
                                    open_class.close()
                            else:
                                # If section folder does not exist
                                section_deleted()
                                # delete the whole institution data base
                        else:
                            _id_.append(name)
                            details1.execute("SELECT rowid FROM COLLEGE WHERE name = ?", _id_)
                            for a1 in details1.fetchall():
                                for rowid1 in a1:
                                    details1.execute(
                                        "DELETE FROM COLLEGE WHERE rowid= " + str(rowid1))
                            else:
                                # If the single institution data base is deleted.
                                for i1 in range(1):
                                    if os.path.exists(section_folder):
                                        for d in os.listdir(section_folder):
                                            if name in d:
                                                dic = os.path.join(section_folder, d)
                                                if dic.endswith(".db"):
                                                    if os.path.exists(dic):
                                                        os.remove(dic)
                                    else:
                                        # If section folder does not exist
                                        section_deleted()
                                        # Delete the whole institution data base
                                else:
                                    for i in range(1):
                                        while True:
                                            check_exist = os.path.join(excel_folder, name)
                                            if os.path.exists(excel_folder) and os.path.exists(check_exist):
                                                temp_folder = os.path.join(parent_dir, '_temp_')
                                                try:
                                                    os.rename(excel_folder, temp_folder)
                                                except OSError:
                                                    root1 = Tk()
                                                    root1.withdraw()
                                                    messagebox.showinfo("Reminder",
                                                                        'Kindly close your Folder')
                                                    root1.destroy()
                                                    continue
                                                else:
                                                    os.rename(temp_folder, excel_folder)
                                                    break
                                            else:
                                                break
                                    else:
                                        institute = os.path.join(excel_folder, name)
                                        if os.path.exists(institute):
                                            for section in os.listdir(institute):
                                                original_section = os.path.join(institute, section)
                                                if '#Deleted' not in section:
                                                    os.rename(original_section, os.path.join(institute,
                                                                                             f'{section} #Deleted '
                                                                                             f'{random.random()}'))
                                            else:
                                                os.rename(institute, os.path.join(excel_folder,
                                                                                  f'{name} #Deleted {random.random()}'))
                else:
                    from_board.commit()
                    from_board.close()
            # if class_folder is not present
            else:
                from_board.commit()
                from_board.close()
                # This function is used to delete the data from the board data base
                college_delete()
                # After function we need to clear section data base from the section folder
                if os.path.exists(section_folder):
                    for z1 in os.listdir(section_folder):
                        if z1.endswith(".db"):
                            section_close = os.path.join(section_folder, z1)
                            os.remove(section_close)

        else:
            # If college_db_folder is DELETED
            for folder in [class_folder, section_folder]:
                if os.path.exists(folder):
                    for file in os.listdir(folder):
                        if file.endswith(".db"):
                            total_close1 = os.path.join(folder, file)
                            if os.path.exists(total_close1):
                                os.remove(total_close1)
            else:
                excel_institution()
    else:
        if os.path.exists(college_db_folder) is False:
            os.mkdir(college_db_folder)
        if os.path.exists(class_folder) is False:
            os.mkdir(class_folder)
        if os.path.exists(section_folder) is False:
            os.mkdir(section_folder)

#  =============================================== EXCEL DATA ==================================#


def excel_file_check(section, institute):
    institution = os.path.join(excel_folder, institute)
    section_inside = os.path.join(institution, section)
    sub_institution = os.path.join(section_inside, f"{section} {str(last_year)} - {str(year)}")
    return sub_institution


def extra_data(period1, n1, sheet1, value, colour_fill):
    for she1 in period1:
        if sheet1[f'{she1}{n1}'].value not in [None]:
            sheet1[f'{she1}{n1}'].value = value
            sheet1[f'{she1}{n1}'].fill = colour_fill
            sheet1[f'{she1}{n1}'].alignment = Alignment(horizontal='center', wrap_text=True)


def excel_maker(institute_name, class_name):
    institution = os.path.join(excel_folder, institute_name)
    extra_folder = os.path.join(institution, class_name)
    sub_institution = os.path.join(extra_folder, f"{class_name} {str(last_year)} - {str(year)}")
    if os.path.exists(excel_folder) is False:
        os.mkdir(excel_folder)
    if os.path.exists(institution) is False:
        os.mkdir(institution)
    if os.path.exists(extra_folder) is False:
        os.mkdir(extra_folder)
    if os.path.exists(sub_institution) is False:
        os.mkdir(sub_institution)
    return sub_institution


def sheet_data(sheet, name, full_column):
    for row, status in enumerate(name, 2):
        if status == 'DELETED':
            sheet[f'A{row}'].value = row - 1
            sheet[f'A{row}'].fill = orange
            sheet[f'A{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
            sheet[f'B{row}'].value = 'DELETED'
            sheet[f'B{row}'].fill = orange
            sheet[f'B{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
            extra_data(full_column, row, sheet, 'DELETED', orange)
        elif status == 'NONE':
            sheet[f'A{row}'].value = row - 1
            sheet[f'A{row}'].fill = yellow
            sheet[f'A{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
            sheet[f'B{row}'].value = 'LEFT INSTITUTION'
            sheet[f'B{row}'].fill = yellow
            sheet[f'B{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
            extra_data(full_column, row, sheet, 'NONE', yellow)
        else:
            sheet[f'A{row}'].value = row - 1
            sheet[f'A{row}'].fill = white
            sheet[f'A{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
            sheet[f'B{row}'].value = status
            sheet[f'B{row}'].fill = white
            sheet[f'B{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
            for column in full_column:
                if sheet[f'{column}{row}'].value in ['NONE', 'DELETED']:
                    sheet[f'{column}{row}'].value = ''
                    sheet[f'{column}{row}'].alignment = Alignment(horizontal='center', wrap_text=True)
                    sheet[f'{column}{row}'].fill = white


def excel_sheet(_section, book_sections, first_name, get_date):
    def sheet2():
        sheet = out_work_book.create_sheet(f'{first_name} {get_date} {month[days.month]}')
        sheet1 = sheet.sheet_properties
        sheet1.tabColor = "FF0000"
        sheet.column_dimensions['B'].width = 30
        for column in table:
            sheet.column_dimensions[column].width = 16
        else:
            sheet.column_dimensions['M'].width = 16
            sheet.column_dimensions['N'].width = 16
            fill1 = sheet.sheet_format
            fill1.customHeight = True
            fill1.defaultRowHeight = float(23)
            fill1.outlineLevelCol = True
            fill1.outlineLevelRow = True
            #sheet.merge_cells('O17:R17')

            out_work_book.save(book_sections)
            wb1 = load_workbook(book_sections)
            sheet1 = wb1[f'{first_name} {get_date} {month[days.month]}']
            for roll1, d in enumerate(hour_data(_section)):
                sheet1[f'{d}{1}'].value = time[roll1]
                sheet1[f'{d}{1}'].alignment = Alignment(vertical='center', horizontal='center')
                sheet1[f'{d}{1}'].border = style
                sheet1[f'{d}{1}'].font = Font(color='FFFFFF', bold=True)
                sheet1[f'{d}{1}'].fill = green
            else:
                name_set = ['Roll No', 'Total Names']
                for b, a in enumerate(['A1', 'B1']):
                    sheet1[a].value = name_set[b]
                    sheet1[a].alignment = Alignment(vertical='center', horizontal='center')
                    sheet1[a].border = style
                    sheet1[a].font = Font(color='FFFFFF', bold=True)
                    sheet1[a].fill = green
                else:
                    wb1.save(book_sections)
    if os.path.exists(book_sections):
        out_work_book = load_workbook(book_sections)
        sheet2()
    else:
        out_work_book = Workbook(book_sections)
        sheet2()
